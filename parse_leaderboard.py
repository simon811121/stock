from bs4 import BeautifulSoup
import requests
from requests.exceptions import ConnectionError
from requests.exceptions import ReadTimeout
import time
import datetime
from datetime import date, timedelta
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import numpy as np
import urllib

# ------------------------------
#    assert funtion
# ------------------------------
def assertFunc(condition, txt, place):
    if not condition:
        print(txt + str(place))
        while(1):
            time.sleep(1)
#----------------------------------------------------------------------
#    parsing data from web
#----------------------------------------------------------------------
# ------------------------------
# <INPUT>
#        url:        for parsing
# <OUTPUT>
#        DataFrame
# ------------------------------
def parseLeaderBoard(url):
    # delay
    time.sleep(5)
    
    # get html info
    html = urllib.request.urlopen(url).read()

    # parse html
    soup = BeautifulSoup(html, 'html5lib')

    # get table
    table = soup.find('table', {'class': 't01'})

    # parsing contents from leaderborad table
    table1 = table.find_all('tr')
    contents = []
    for row in table1[2:]:      
        row = row.find_all('td')
        title = []
        for row1 in row: 
            col = row1.get_text()
            title.append(col)
        contents.append(title)

    # form DataFrame data from parsing result data
    df = pd.DataFrame.from_dict(contents[1:])
    contents[0][1] = table1[1].find_all('td')[0].text + contents[0][1]
    contents[0][2] = table1[1].find_all('td')[0].text + '張數'
    contents[0][6] = table1[1].find_all('td')[1].text + contents[0][6]
    contents[0][7] = table1[1].find_all('td')[1].text + '張數'
    df.columns = contents[0]
    date = []
    for i in range(0,len(df)):
        date.append(table1[0].find("div").get_text()[3:])
    df['法人買賣日期'] = pd.DataFrame.from_dict(date)
    df.index = df['名次']
    del df['名次']
    
    return df

#----------------------------------------------------------------------
#    save data
#----------------------------------------------------------------------
# ------------------------------
# <INPUT>
#        saveDf:    dataframe to save
#        startCol:  excel column offset to save
#        startRow:  excel row offset to save
#        txt:       title txt to save
# <OUTPUT>
#        saved excel in directory
# ------------------------------
def saveToExcel(saveDf, startCol, startRow, txt):
    if len(saveDf) == 0:
        print(txt + ' has no data')
        return None
    try:
        dateString = saveDf['法人買賣日期']
        dateString1 = dateString.values[0].split("/")
    except Exception as error:
        print(error)
        assertFunc(0, 'error code logic', 1)

    year = str(date.today().year)
    if date.today().month < 10:
        month = ' 0' + str(date.today().month)
    else:
        month = ' ' + str(date.today().month)

    outputFileName = year + month + '月 三大法人買賣超 紀錄.xlsx'
    sheetName = dateString1[0] + '-' + dateString1[1]
    titleRow = 1 if ((startRow - 1) <= 0) else (startRow - 1)
    titleCol = 1 if ((startCol - 1) <= 0) else (startCol - 1)

    try:
        outputFile = load_workbook(outputFileName)
        writer = pd.ExcelWriter(outputFileName, engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        writer.book = outputFile
        writer.sheets = dict((ws.title, ws) for ws in outputFile.worksheets)
        saveDf.to_excel(writer, sheet_name=sheetName, startcol=startCol, startrow=startRow)
        worksheet = writer.sheets[sheetName]
        worksheet.cell(row=titleRow, column=titleCol).value = txt
        writer.save()
    except Exception as error: 
        print(error)
        saveDf.to_excel(outputFileName, sheet_name=sheetName, startcol=startCol, startrow=startRow)
        outputFile = load_workbook(outputFileName)
        worksheet = outputFile.get_sheet_by_name(sheetName)
        worksheet.cell(row=titleRow, column=titleCol).value = txt
        outputFile.save(outputFileName)

#----------------------------------------------------------------------
#    compare data
#----------------------------------------------------------------------
# ------------------------------
# <INPUT>
#        df1:    dataframe 1st to compare
#        df2:    dataframe 2nd to compare
#        startCol:  excel column offset to save
#        startRow:  excel row offset to save
#        txt:       title txt to save
# <OUTPUT>
#        result df
# --------------------------------------
# 這個 function 是用來取得剛爬完蟲的資料
# --------------------------------------
def getRepeatStockIdDf(df1, df2, startCol, startRow, txt):
    rsltDfBuy = pd.merge(df1, df2, on=['買超股票名稱'], how='inner')
    rsltDfSell = pd.merge(df1, df2, on=['賣超股票名稱'], how='inner')
    if len(rsltDfBuy) == 0 and len(rsltDfSell) == 0: # no repeate stock id between two rank list
        return rsltDfBuy, rsltDfSell
    rsltDfBuy = rsltDfBuy[['買超股票名稱', '法人買賣日期_x']]
    rsltDfBuy.dropna(subset=["買超股票名稱"], inplace=True)
    rsltDfSell = rsltDfSell[['賣超股票名稱', '法人買賣日期_x']]
    rsltDfSell.dropna(subset=["賣超股票名稱"], inplace=True)
    rsltDfBuy.rename(columns={'法人買賣日期_x':'法人買賣日期'}, inplace=True)
    rsltDfSell.rename(columns={'法人買賣日期_x':'法人買賣日期'}, inplace=True)
    rsltDfBuy.index = np.arange(1, len(rsltDfBuy) + 1)
    rsltDfSell.index = np.arange(1, len(rsltDfSell) + 1)
    RowOfst = 5
    if len(rsltDfBuy) != 0:
        saveToExcel(rsltDfBuy, startCol, startRow, txt + '買超')
        RowOfst = len(rsltDfBuy)
    if len(rsltDfSell) != 0:
        saveToExcel(rsltDfSell, startCol, startRow + RowOfst + 4, txt + '賣超')
    return rsltDfBuy, rsltDfSell, startCol + len(rsltDfBuy.columns) + 4
# ------------------------------
# <INPUT>
#        df1:    dataframe 1st to compare
#        df2:    dataframe 2nd to compare
# <OUTPUT>
#        result df
# ---------------------------------------------------
# 這個 function 是用來比較今日爬完結果，及昨日爬完結果
# ---------------------------------------------------
def getRepeatStockRankDf(df1, df2, txt):
    if len(df1) == 0 or len(df2) == 0:
        return pd.DataFrame()
    try:        
        rsltDf = pd.merge(df1, df2, on=[txt+'超股票名稱', '法人買賣日期'], how='inner')
    except:
        pass

    return rsltDf

#----------------------------------------------------------------------
#    get yesterday's rank overlap detail
#----------------------------------------------------------------------
#----------------------------------------------------------------------
#    get yesterday's data
#----------------------------------------------------------------------
# ------------------------------
# <INPUT>
#        excel_name:    excel 名稱
#        sheet_name:    工作表名稱
#        read_ofst:     從哪開始讀
#        txt:           過濾文字
# <OUTPUT>
#        y_rank_df_buy:   法人買超股票 dataframe
#        y_rank_df_sell:  法人賣超股票 dataframe
# ------------------------------
def getYestData(excel_name, sheet_name, read_ofst, txt):
    # y = yesterday; [read_ofst, read_ofst + 1] = 代號, 名稱
    try:
        y_rank_df_buy = pd.read_excel(excel_name, sheet_name=sheet_name, nrows=50, usecols=[read_ofst, read_ofst + 1])
        y_rank_df_buy = y_rank_df_buy.dropna()
        y_rank_df_buy.index = np.arange(1, len(y_rank_df_buy) + 1)
        y_rank_df_buy = y_rank_df_buy.rename(columns={'Unnamed: '+str(read_ofst):'買超股票名稱', 'Unnamed: '+str(read_ofst+1):'法人買賣日期'})
        y_rank_df_buy = y_rank_df_buy.rename(columns={'買超股票名稱'+txt:'買超股票名稱', '法人買賣日期'+txt: '法人買賣日期'})
        try:
            y_rank_df_sell = y_rank_df_buy[y_rank_df_buy.index[y_rank_df_buy['買超股票名稱']=='賣超股票名稱'][0]:]
            y_rank_df_sell.rename(columns={'買超股票名稱':'賣超股票名稱'},inplace=True)
            y_rank_df_sell.index = np.arange(1, len(y_rank_df_sell) + 1)
            y_rank_df_buy = y_rank_df_buy[:y_rank_df_buy.index[y_rank_df_buy['買超股票名稱']=='賣超股票名稱'][0] - 1]
        except:
            y_rank_df_sell = pd.DataFrame()
            print(txt + " 沒有賣超股票")
    except:
        y_rank_df_buy = pd.DataFrame()
        y_rank_df_sell = pd.DataFrame()
    
    return y_rank_df_buy, y_rank_df_sell
#----------------------------------------------------------------------
#    code main
#----------------------------------------------------------------------
print(datetime.datetime.now())
LEADERBOARD_MAX_RANK_CHK = 50

# a = amount = 成交量
# TODO:
# 成交量與昨日排行比
# url = 'https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E6%88%90%E4%BA%A4%E9%87%8F%E5%A2%9E%E5%8A%A0%E5%BC%B5%E6%95%B8%E2%80%93%E7%95%B6%E6%97%A5%E6%88%90%E4%BA%A4%E9%87%8F%E8%88%87%E6%98%A8%E6%97%A5%E6%AF%94%40%40%E6%88%90%E4%BA%A4%E9%87%8F%E5%A2%9E%E5%8A%A0%E5%BC%B5%E6%95%B8%40%40%E7%95%B6%E6%97%A5%E6%88%90%E4%BA%A4%E9%87%8F%E8%88%87%E6%98%A8%E6%97%A5%E6%AF%94'
# url = 'https://fubon-ebrokerdj.fbs.com.tw/Z/ZG/ZG_B.djhtm' # 上市量增
# aRstDf = parseLeaderBoard(url, LEADERBOARD_MAX_RANK_CHK)
# aRstDf = aRstDf[["代號", "名稱"]]
#aRstDf2 = pd.concat([aRstDf, fRsltDf['法人買賣日期']],axis=1)
#saveToExcel(aRstDf2, 0, 1, '成交量與昨日排行比')

# f = foreign = 外資
# i = invset trust = 投信
# d = dealer = 自營商
# 外資單日
url = 'https://fubon-ebrokerdj.fbs.com.tw/Z/ZG/ZGK_D.djhtm'
fRsltDf = parseLeaderBoard(url)
saveToExcel(fRsltDf, 0, (len(fRsltDf) + 5), '外資 單日')

# 投信單日
url = 'https://fubon-ebrokerdj.fbs.com.tw/Z/ZG/ZGK_DD.djhtm'
iRsltDf = parseLeaderBoard(url)
saveToExcel(iRsltDf, 0, ((len(iRsltDf) + 5) * 2), '投信 單日')

# 自營商單日
url = 'https://fubon-ebrokerdj.fbs.com.tw/Z/ZG/ZGK_DB.djhtm'
dRsltDf = parseLeaderBoard(url)
saveToExcel(dRsltDf, 0, ((len(dRsltDf) + 5) * 3), '自營商 單日')

# 外資 & 投信 單日
fileColumnOfst = 1
f_i_RsltDf_buy, f_i_RsltDf_sell, fileColumnOfst = getRepeatStockIdDf(fRsltDf, iRsltDf, fileColumnOfst, 0, '外資 & 投信 單日')

# 投信 & 自營商 單日
i_d_RsltDf_buy, i_d_RsltDf_sell, fileColumnOfst = getRepeatStockIdDf(iRsltDf, dRsltDf, fileColumnOfst, 0, '投信 & 自營商 單日')

# 外資 & 自營商 單日
f_d_RsltDf_buy, f_d_RsltDf_sell, fileColumnOfst = getRepeatStockIdDf(fRsltDf, dRsltDf, fileColumnOfst, 0, '外資 & 自營商 單日')

# 外資 & 投信 & 自營商 單日
f_i_RsltDf = pd.concat([f_i_RsltDf_buy, f_i_RsltDf_sell], axis=0)
i_d_RsltDf = pd.concat([i_d_RsltDf_buy, i_d_RsltDf_sell], axis=0)
f_i_d_RsltDf_buy, f_i_d_RsltDf_sell, fileColumnOfst = getRepeatStockIdDf(f_i_RsltDf, i_d_RsltDf, fileColumnOfst, 0, '外資 & 投信 & 自營商 單日')

# 計算連續買超儲存 colunm 排數
LEADERBOARD_OVERLAP_DATA_COL_OFST = fileColumnOfst

# 設定國定假日
holidays_array_valid = [1, # 2020
                        1, # 2021
                        0, # 2022
                        0, # 2023
                        0, # 2024
                        0] # 2025
holidays_in_2020_month = [10, 10, 10, 1] # month
holidays_in_2020_day   = [ 9,  2,  1, 1] # day, # 必須從後面的往回填
holidays_in_2021_month = [12, 10,  9,  9,  6, 5,  4, 4, 4, 3,  2,  2,  2,  2,  2, 2, 2, 1] # month
holidays_in_2021_day   = [31, 11, 21, 20, 14, 7, 30, 5, 2, 1, 16, 15, 12, 11, 10, 9, 8, 1] # day, # 必須從後面的往回填
holidays_len = [4, # 2020
                18, # 2021
                0, # 2022
                0, # 2023
                0, # 2024
                0] # 2025

yesterday = date.today() - timedelta(days=1)
if not holidays_array_valid[yesterday.year - 2020]:
    print('please fill holidays in ' + str(yesterday.year))
    assertFunc(0, 'error code logic', 4)
if not holidays_len[yesterday.year - 2020]:
    print('please fill holidays length in ' + str(yesterday.year))
    assertFunc(0, 'error code logic', 5)

# 取得昨天的排行，並得到任兩大法人連續
yesterdat_is_holiday = False
i = 0
while (i < holidays_len[yesterday.year - 2020]):
    holidays = date(yesterday.year, holidays_in_2021_month[i], holidays_in_2021_day[i])
    if yesterday == holidays:
        yesterdat_is_holiday = True
        break
    if yesterday.month > holidays_in_2021_month[i]:
        break
    i+= 1

while ((yesterday.weekday() >= 5) or (yesterdat_is_holiday)):  # 只挑 1 ~ 5
    yesterdat_is_holiday = False
    yesterday = yesterday - timedelta(days=1)
    i = 0
    while (i < holidays_len[yesterday.year - 2020]):
        holidays = date(yesterday.year, holidays_in_2021_month[i], holidays_in_2021_day[i])
        if yesterday == holidays:
            yesterday = yesterday - timedelta(days=1)
        if yesterday.month > holidays_in_2021_month[i]:
            break
        i += 1
        

if yesterday.month < 10:
    excel_name = str(yesterday.year) + ' 0' + str(yesterday.month) + '月 三大法人買賣超 紀錄.xlsx'
else:
    excel_name = str(yesterday.year) + ' ' + str(yesterday.month) + '月 三大法人買賣超 紀錄.xlsx'

month_str = '0' if yesterday.month < 10 else ''
if yesterday.day < 10:
    sheet_name = month_str + str(yesterday.month) + '-0' + str(yesterday.day)
else:
    sheet_name = month_str + str(yesterday.month) + '-' + str(yesterday.day)

# 取得昨天的排行榜，計算連續多日排行榜
y_f_i_RsltDf_buy, y_f_i_RsltDf_sell = getYestData(excel_name, sheet_name, 2, '')
y_i_d_RsltDf_buy, y_i_d_RsltDf_sell = getYestData(excel_name, sheet_name, 8, '.1')
y_f_d_RsltDf_buy, y_f_d_RsltDf_sell = getYestData(excel_name, sheet_name, 14, '.2')
y_f_i_d_RsltDf_buy, y_f_i_d_RsltDf_sell = getYestData(excel_name, sheet_name, 20, '.3')

# 找連續
add_col = 4
fileColOfst = LEADERBOARD_OVERLAP_DATA_COL_OFST
fileRowOfst = 5
# 買超
c_rank_Df_buy = None
try:
    df1 = getRepeatStockRankDf(y_f_i_d_RsltDf_buy, f_i_RsltDf_buy, '買')
    df2 = getRepeatStockRankDf(y_f_i_d_RsltDf_buy, i_d_RsltDf_buy, '買')
    df3 = getRepeatStockRankDf(y_f_i_d_RsltDf_buy, f_d_RsltDf_buy, '買')
    df4 = getRepeatStockRankDf(y_f_i_d_RsltDf_buy, f_i_d_RsltDf_buy, '買')
    tmp_Df = [df1, df2, df3, df4]
    c_rank_Df_buy = pd.concat(tmp_Df)
    c_rank_Df_buy = c_rank_Df_buy.drop_duplicates()
    c_rank_Df_buy.index = np.arange(1, len(c_rank_Df_buy) + 1)

    saveToExcel(c_rank_Df_buy, fileColOfst, 0, '任兩大法人連續買超')
    add_col = len(c_rank_Df_buy.columns) + 4
    fileRowOfst = len(c_rank_Df_buy)
except Exception as error:
    print(error)
    assertFunc(0, 'error code logic', 6)

# 賣超
c_rank_Df_sell = None
try:
    df1 = getRepeatStockRankDf(y_f_i_d_RsltDf_sell, f_i_RsltDf_sell, '賣')
    df2 = getRepeatStockRankDf(y_f_i_d_RsltDf_sell, i_d_RsltDf_sell, '賣')
    df3 = getRepeatStockRankDf(y_f_i_d_RsltDf_sell, f_d_RsltDf_sell, '賣')
    df4 = getRepeatStockRankDf(y_f_i_d_RsltDf_sell, f_i_d_RsltDf_sell, '賣')
    tmp_Df = [df1, df2, df3, df4]
    c_rank_Df_sell = pd.concat(tmp_Df)
    c_rank_Df_sell = c_rank_Df_sell.drop_duplicates()
    c_rank_Df_sell.index = np.arange(1, len(c_rank_Df_sell) + 1)
    
    saveToExcel(c_rank_Df_sell, fileColOfst, fileRowOfst + 4, '任兩大法人連續賣超')
    add_col = len(c_rank_Df_sell.columns) + 4
except Exception as error:
    print(error)
    assertFunc(0, 'error code logic', 7)

fileColOfst += (add_col + 2)

# 取得要輸入的股票(買超)
tod_Df_buy = [f_i_RsltDf_buy, i_d_RsltDf_buy, f_d_RsltDf_buy]
s_rank_Df_buy = pd.concat(tod_Df_buy)
tod_Df_buy = [s_rank_Df_buy, f_i_d_RsltDf_buy, c_rank_Df_buy]
s_rank_Df_buy = pd.concat(tod_Df_buy)
s_rank_Df_buy = s_rank_Df_buy.drop_duplicates(keep=False)
s_rank_Df_buy.index = np.arange(1, len(s_rank_Df_buy) + 1)
fileRowOfst = 5
if len(s_rank_Df_buy):
    fileRowOfst = len(s_rank_Df_buy)
saveToExcel(s_rank_Df_buy, fileColOfst, 0, '單日該輸入的股票(買超)')

# 取得要輸入的股票(買超)
tod_Df_sell = [f_i_RsltDf_sell, i_d_RsltDf_sell, f_d_RsltDf_sell]
s_rank_Df_sell = pd.concat(tod_Df_sell)
tod_Df_sell = [s_rank_Df_sell, f_i_d_RsltDf_sell, c_rank_Df_sell]
s_rank_Df_sell = pd.concat(tod_Df_sell)
s_rank_Df_sell = s_rank_Df_sell.drop_duplicates(keep=False)
s_rank_Df_sell.index = np.arange(1, len(s_rank_Df_sell) + 1)
saveToExcel(s_rank_Df_sell, fileColOfst, fileRowOfst + 4, '單日該輸入的股票(賣超)')

print("code end")